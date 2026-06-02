#!/usr/bin/env python3
"""
Yandex Wordstat API client (native API: https://api.wordstat.yandex.net).

Auth: OAuth token passed as `Authorization: Bearer <token>`.
Get the token via the steps in ../SKILL.md, then export it:

    export WORDSTAT_OAUTH_TOKEN="y0_AgAAA..."

Methods implemented (per https://yandex.com/support2/wordstat/ru/content/api-structure):
  - topRequests   POST /v1/topRequests   популярные + похожие запросы (за 30 дней)
  - dynamics      POST /v1/dynamics      динамика частотности во времени
  - regions       POST /v1/regions       распределение по регионам (+ affinityIndex)
  - regionsTree   POST /v1/getRegionsTree дерево/список регионов
  - userInfo      POST /v1/userInfo      остаток квоты

Limits (default personal quota): 10 req/sec, 1000 req/day.
On 429 the API returns "Time to refill: N seconds" — the client backs off and retries.

CLI examples:
  python wordstat.py top "купить кроссовки" --regions 213 --devices phone
  python wordstat.py top "ремонт квартир" --num 200 --csv out.csv
  python wordstat.py dynamics "купить елку" --period monthly
  python wordstat.py regions "доставка пиццы"
  python wordstat.py regions-tree
  python wordstat.py quota
  python wordstat.py core phrases.txt --regions 225 --xlsx seo_core.xlsx

`core` runs topRequests for every phrase in a file (one per line) and aggregates
all collected phrases+counts into a single deduplicated table — ready for an
SEO core or a Yandex Direct keyword sheet.
"""

import argparse
import csv
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request

BASE_URL = "https://api.wordstat.yandex.net"
TOKEN_ENV = "WORDSTAT_OAUTH_TOKEN"
# Бренд-строка выводится в stderr (не ломает JSON в stdout).
BRAND = "— Powered by leadl.ai · https://leadl.ai —"

# Папка для авто-выгрузки результатов поиска в Markdown (рядом со скиллом: ../results).
RESULTS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "results"
)

# Conservative client-side throttle (default quota = 10 req/sec).
MIN_INTERVAL_SEC = 0.15
MAX_RETRIES = 5

_last_call = [0.0]


class WordstatError(Exception):
    pass


def _token():
    tok = os.environ.get(TOKEN_ENV, "").strip()
    if not tok:
        raise WordstatError(
            f"Не найден OAuth-токен. Установи переменную окружения {TOKEN_ENV}.\n"
            f"Как получить токен — см. SKILL.md."
        )
    return tok


def _throttle():
    elapsed = time.time() - _last_call[0]
    if elapsed < MIN_INTERVAL_SEC:
        time.sleep(MIN_INTERVAL_SEC - elapsed)
    _last_call[0] = time.time()


def call(method: str, payload: dict) -> dict:
    """POST to /v1/<method> with the JSON payload; return parsed JSON."""
    url = f"{BASE_URL}/v1/{method}"
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    for attempt in range(1, MAX_RETRIES + 1):
        _throttle()
        req = urllib.request.Request(url, data=body, method="POST")
        req.add_header("Content-Type", "application/json; charset=utf-8")
        req.add_header("Authorization", f"Bearer {_token()}")
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                raw = resp.read().decode("utf-8")
                return json.loads(raw) if raw else {}
        except urllib.error.HTTPError as e:
            detail = e.read().decode("utf-8", "replace")
            # 429 = persistent quota; 503 = global quota — both are retryable.
            if e.code in (429, 503) and attempt < MAX_RETRIES:
                wait = _refill_seconds(detail, default=2 * attempt)
                sys.stderr.write(
                    f"[{e.code}] квота превышена, жду {wait}s "
                    f"(попытка {attempt}/{MAX_RETRIES})\n"
                )
                time.sleep(wait)
                continue
            raise WordstatError(f"HTTP {e.code} от {method}: {detail}") from e
        except urllib.error.URLError as e:
            if attempt < MAX_RETRIES:
                time.sleep(2 * attempt)
                continue
            raise WordstatError(f"Сетевая ошибка при вызове {method}: {e}") from e
    raise WordstatError(f"Не удалось получить ответ от {method} за {MAX_RETRIES} попыток")


def _refill_seconds(detail: str, default: int) -> int:
    """Parse 'Time to refill: N seconds' if present."""
    import re

    m = re.search(r"refill:\s*(\d+)", detail)
    return int(m.group(1)) if m else default


# ---------------------------------------------------------------------------
# API method wrappers
# ---------------------------------------------------------------------------

def top_requests(phrase, regions=None, devices=None, num_phrases=None):
    payload = {"phrase": phrase}
    if regions:
        payload["regions"] = regions
    if devices:
        payload["devices"] = devices
    if num_phrases:
        payload["numPhrases"] = num_phrases
    return call("topRequests", payload)


def _default_range(period):
    """Дефолтное окно дат (YYYY-MM-DD) под период, если --from/--to не заданы.

    API требует выравнивания: monthly -> первый день месяца,
    weekly -> понедельник. Для daily ограничений нет.
    """
    from datetime import date, timedelta

    today = date.today()
    if period == "monthly":
        # последний день прошлого месяца … первый день 12 месяцами раньше
        to_d = today.replace(day=1) - timedelta(days=1)
        y, m = to_d.year, to_d.month - 11
        while m <= 0:
            m += 12
            y -= 1
        return date(y, m, 1).isoformat(), to_d.isoformat()
    if period == "weekly":
        # неделя Пн…Вс: toDate — ближайшее воскресенье, fromDate — понедельник 26 нед. назад
        days_since_sun = (today.weekday() - 6) % 7  # Пн=0 … Вс=6
        to_d = today - timedelta(days=days_since_sun)
        from_d = to_d - timedelta(weeks=26) + timedelta(days=1)
        return from_d.isoformat(), to_d.isoformat()
    return (today - timedelta(days=60)).isoformat(), today.isoformat()


def dynamics(phrase, period="monthly", regions=None, devices=None,
             from_date=None, to_date=None):
    # API /v1/dynamics требует fromDate/toDate в формате YYYY-MM-DD.
    payload = {"phrase": phrase, "period": period}
    if from_date:
        payload["fromDate"] = from_date
    if to_date:
        payload["toDate"] = to_date
    if regions:
        payload["regions"] = regions
    if devices:
        payload["devices"] = devices
    return call("dynamics", payload)


def regions(phrase, devices=None):
    payload = {"phrase": phrase}
    if devices:
        payload["devices"] = devices
    return call("regions", payload)


def regions_tree():
    return call("getRegionsTree", {})


def user_info():
    return call("userInfo", {})


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------

def write_csv(rows, path, fieldnames):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    sys.stderr.write(f"CSV сохранён: {path} ({len(rows)} строк)\n")


def write_xlsx(rows, path, fieldnames, sheet="Wordstat"):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise WordstatError(
            "Для --xlsx нужен openpyxl: pip install openpyxl --break-system-packages"
        )
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(fieldnames)
    for r in rows:
        ws.append([r.get(k, "") for k in fieldnames])
    wb.save(path)
    sys.stderr.write(f"XLSX сохранён: {path} ({len(rows)} строк)\n")


def _slug(text):
    """Имя-слаг из тела запроса: убрать операторы Вордстата, нелат/некир -> дефис."""
    s = (text or "").strip().lower()
    s = re.sub(r'["\'\[\]!+]', "", s)          # операторы вордстата и кавычки
    s = re.sub(r"[^\w]+", "-", s, flags=re.U)  # \w c re.U сохраняет кириллицу
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s[:80] or "wordstat"


def write_xlsx_sheets(path, sheets):
    """Записать несколько листов в ОДНУ книгу. sheets: список (title, fieldnames, rows)."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise WordstatError(
            "Для xlsx нужен openpyxl: pip install openpyxl --break-system-packages"
        )
    wb = Workbook()
    for i, (title, fieldnames, rows) in enumerate(sheets):
        safe = re.sub(r"[:\\/?*\[\]]", " ", str(title))[:31] or "sheet"
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = safe
        ws.append(list(fieldnames))
        for r in rows:
            ws.append([r.get(k, "") for k in fieldnames])
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    names = ", ".join(t for t, _, _ in sheets)
    total = sum(len(r) for _, _, r in sheets)
    sys.stderr.write(f"XLSX сохранён: {path} (листы: {names}; всего строк {total})\n")


def emit(rows, fieldnames, args):
    if getattr(args, "csv", None):
        write_csv(rows, args.csv, fieldnames)
    if getattr(args, "xlsx", None):
        write_xlsx(rows, args.xlsx, fieldnames)
    if not getattr(args, "csv", None) and not getattr(args, "xlsx", None):
        print(json.dumps(rows, ensure_ascii=False, indent=2))


def _parse_top(data):
    """Normalize topRequests response into flat rows."""
    rows = []
    seed = data.get("requestPhrase") or data.get("phrase") or ""
    total = data.get("totalCount") or data.get("count")
    for item in data.get("topRequests", []) or []:
        rows.append({
            "seed": seed,
            "type": "top",
            "phrase": item.get("phrase", ""),
            "count": item.get("count", ""),
            "seed_total": total,
        })
    # "похожие" / ассоциации могут приходить отдельным массивом
    for key in ("associations", "searchedWith", "similar"):
        for item in data.get(key, []) or []:
            rows.append({
                "seed": seed,
                "type": "similar",
                "phrase": item.get("phrase", ""),
                "count": item.get("count", ""),
                "seed_total": total,
            })
    return rows


# ---------------------------------------------------------------------------
# CLI command handlers
# ---------------------------------------------------------------------------

def cmd_top(args):
    data = top_requests(args.phrase, args.regions, args.devices, args.num)
    rows = _parse_top(data)
    emit(rows, ["seed", "type", "phrase", "count", "seed_total"], args)


def cmd_dynamics(args):
    frm, to = args.from_date, args.to_date
    if not frm or not to:
        frm, to = _default_range(args.period)
    sys.stderr.write(f"Период {args.period}: {frm} … {to}\n")
    data = dynamics(args.phrase, args.period, args.regions, args.devices, frm, to)
    rows = [
        {
            "phrase": args.phrase,
            "date": d.get("date", ""),
            "count": d.get("count", ""),
            "share": d.get("share", ""),
        }
        for d in data.get("dynamics", []) or []
    ]
    emit(rows, ["phrase", "date", "count", "share"], args)


def cmd_regions(args):
    data = regions(args.phrase, args.devices)
    rows = [
        {
            "phrase": args.phrase,
            "regionId": r.get("regionId", ""),
            "count": r.get("count", ""),
            "share": r.get("share", ""),
            "affinityIndex": r.get("affinityIndex", ""),
        }
        for r in data.get("regions", []) or []
    ]
    emit(rows, ["phrase", "regionId", "count", "share", "affinityIndex"], args)


def cmd_regions_tree(args):
    data = regions_tree()
    print(json.dumps(data, ensure_ascii=False, indent=2))


def cmd_quota(args):
    print(json.dumps(user_info(), ensure_ascii=False, indent=2))


# Мелкий набор служебных слов — не считаем их «корнями» при оценке релевантности.
RU_STOP = {
    "для", "под", "что", "это", "как", "или", "при", "без", "про", "над",
    "так", "там", "все", "всё", "его", "чем", "где", "кто", "той", "ключ",
    "такое", "такой", "простыми", "словами", "значит",
}


def _words(text):
    return re.findall(r"[a-zа-яё0-9]+", (text or "").lower())


def _common_prefix(a, b):
    n = 0
    for ca, cb in zip(a, b):
        if ca != cb:
            break
        n += 1
    return n


def _seed_roots(seeds):
    """Значимые слова из сид-фраз — основа для оценки релевантности."""
    roots = set()
    for s in seeds:
        for w in _words(s):
            if len(w) >= 4 and w not in RU_STOP:
                roots.add(w)
    return roots


def _match_root(phrase_words, roots):
    """Самый специфичный сид-корень, к которому относится фраза (учёт морфологии
    через общий префикс), либо None если фраза не релевантна ни одному сиду."""
    best, best_len = None, 0
    for w in phrase_words:
        if len(w) < 4:
            continue
        for r in roots:
            cp = _common_prefix(w, r)
            if cp >= max(4, len(r) - 3) and cp > best_len:
                best, best_len = r, cp
    return best


def clean_cluster(rows, seeds, min_count=0, stop=None):
    """Обобщённая очистка пула: оставить фразы, релевантные сидам, и присвоить
    кластер = матчнутый сид-корень. Доменно-нейтрально (корни берутся из сидов)."""
    roots = _seed_roots(seeds)
    stop = {w.lower() for w in (stop or [])}
    kept = []
    for r in rows:
        pw = _words(r["phrase"])
        try:
            cnt = int(r["count"])
        except (ValueError, TypeError):
            cnt = 0
        if cnt < min_count:
            continue
        if stop & set(pw):
            continue
        root = _match_root(pw, roots)
        if not root:
            continue
        kept.append({"cluster": root, "phrase": r["phrase"],
                     "count": r["count"], "sources": r["sources"]})

    from collections import Counter
    sizes = Counter(x["cluster"] for x in kept)
    for x in kept:
        if sizes[x["cluster"]] < 3:           # мелкие кластеры -> "прочее"
            x["cluster"] = "прочее"

    def _key(x):
        try:
            return (x["cluster"], -int(x["count"]))
        except (ValueError, TypeError):
            return (x["cluster"], 0)
    kept.sort(key=_key)
    return kept


def cmd_core(args):
    """Batch: topRequests for every seed phrase in a file -> aggregated table."""
    with open(args.file, encoding="utf-8") as f:
        seeds = [ln.strip() for ln in f if ln.strip() and not ln.startswith("#")]
    if not seeds:
        raise WordstatError(f"В файле {args.file} не найдено фраз.")

    agg = {}  # phrase -> {"count": int, "sources": set}
    for i, seed in enumerate(seeds, 1):
        sys.stderr.write(f"[{i}/{len(seeds)}] {seed}\n")
        try:
            data = top_requests(seed, args.regions, args.devices, args.num)
        except WordstatError as e:
            sys.stderr.write(f"  пропуск ({e})\n")
            continue
        for row in _parse_top(data):
            ph = row["phrase"]
            if not ph:
                continue
            entry = agg.setdefault(ph, {"count": row["count"], "sources": set()})
            entry["sources"].add(seed)
            # keep the max numeric count seen
            try:
                if int(row["count"]) > int(entry["count"]):
                    entry["count"] = row["count"]
            except (ValueError, TypeError):
                pass

    rows = [
        {"phrase": ph, "count": v["count"], "sources": ", ".join(sorted(v["sources"]))}
        for ph, v in agg.items()
    ]
    # sort by count desc when numeric
    def _key(r):
        try:
            return -int(r["count"])
        except (ValueError, TypeError):
            return 0
    rows.sort(key=_key)
    sys.stderr.write(f"Собрано уникальных фраз: {len(rows)}\n")
    if getattr(args, "clean", False):
        rows = clean_cluster(rows, seeds, args.min_count or 0, args.stop)
        n_cl = len({r["cluster"] for r in rows})
        sys.stderr.write(f"После очистки: {len(rows)} фраз, {n_cl} кластеров\n")
        emit(rows, ["cluster", "phrase", "count", "sources"], args)
    else:
        emit(rows, ["phrase", "count", "sources"], args)


def _safe_int(v):
    try:
        return int(v)
    except (ValueError, TypeError):
        return 0


def cmd_research(args):
    """Единый отчёт по фразе: top + dynamics + regions -> ОДИН xlsx с вкладками."""
    from datetime import date

    phrase = args.phrase
    sheets = []

    sys.stderr.write(f"[research] {phrase}\n")
    try:
        data = top_requests(phrase, args.regions, args.devices, args.num)
        sheets.append(("top", ["seed", "type", "phrase", "count", "seed_total"],
                       _parse_top(data)))
    except WordstatError as e:
        sys.stderr.write(f"  top пропущен: {e}\n")

    try:
        frm, to = args.from_date, args.to_date
        if not frm or not to:
            frm, to = _default_range(args.period)
        sys.stderr.write(f"  dynamics {args.period}: {frm} … {to}\n")
        d = dynamics(phrase, args.period, args.regions, args.devices, frm, to)
        rows = [{"phrase": phrase, "date": x.get("date", ""),
                 "count": x.get("count", ""), "share": x.get("share", "")}
                for x in d.get("dynamics", []) or []]
        sheets.append(("dynamics", ["phrase", "date", "count", "share"], rows))
    except WordstatError as e:
        sys.stderr.write(f"  dynamics пропущен: {e}\n")

    try:
        d = regions(phrase, args.devices)
        rows = [{"phrase": phrase, "regionId": r.get("regionId", ""),
                 "count": r.get("count", ""), "share": r.get("share", ""),
                 "affinityIndex": r.get("affinityIndex", "")}
                for r in d.get("regions", []) or []]
        rows.sort(key=lambda r: -_safe_int(r["count"]))  # по убыванию спроса
        sheets.append(("regions",
                       ["phrase", "regionId", "count", "share", "affinityIndex"], rows))
    except WordstatError as e:
        sys.stderr.write(f"  regions пропущен: {e}\n")

    if not sheets:
        raise WordstatError("Все запросы research завершились ошибкой — нечего сохранять.")

    path = args.xlsx
    if not path:
        os.makedirs(RESULTS_DIR, exist_ok=True)
        path = os.path.join(RESULTS_DIR, f"{_slug(phrase)}_{date.today().isoformat()}.xlsx")
    write_xlsx_sheets(path, sheets)


def build_parser():
    p = argparse.ArgumentParser(description="Yandex Wordstat API client")
    sub = p.add_subparsers(dest="cmd", required=True)

    def add_common(sp, with_num=False):
        sp.add_argument("--regions", type=int, nargs="*", help="ID регионов, напр. 213 2 225")
        sp.add_argument(
            "--devices", nargs="*",
            choices=["all", "desktop", "phone", "tablet"],
            help="типы устройств",
        )
        if with_num:
            sp.add_argument("--num", type=int, default=None,
                            help="число фраз (по умолч. 50, макс 2000)")
        sp.add_argument("--csv", help="сохранить в CSV")
        sp.add_argument("--xlsx", help="сохранить в XLSX")

    sp = sub.add_parser("top", help="популярные + похожие запросы")
    sp.add_argument("phrase")
    add_common(sp, with_num=True)
    sp.set_defaults(func=cmd_top)

    sp = sub.add_parser("dynamics", help="динамика частотности во времени")
    sp.add_argument("phrase")
    sp.add_argument("--period", default="monthly",
                    choices=["daily", "weekly", "monthly"])
    sp.add_argument("--from", dest="from_date", metavar="YYYY-MM-DD",
                    help="дата начала (по умолч. зависит от --period)")
    sp.add_argument("--to", dest="to_date", metavar="YYYY-MM-DD",
                    help="дата конца (по умолч. сегодня)")
    add_common(sp)
    sp.set_defaults(func=cmd_dynamics)

    sp = sub.add_parser("regions", help="распределение по регионам")
    sp.add_argument("phrase")
    add_common(sp)
    sp.set_defaults(func=cmd_regions)

    sp = sub.add_parser("regions-tree", help="дерево регионов")
    sp.set_defaults(func=cmd_regions_tree)

    sp = sub.add_parser("quota", help="остаток квоты (userInfo)")
    sp.set_defaults(func=cmd_quota)

    sp = sub.add_parser(
        "research",
        help="единый отчёт по фразе (top+dynamics+regions) -> один xlsx с вкладками")
    sp.add_argument("phrase")
    sp.add_argument("--regions", type=int, nargs="*", help="ID регионов, напр. 213 2 225")
    sp.add_argument("--devices", nargs="*",
                    choices=["all", "desktop", "phone", "tablet"], help="типы устройств")
    sp.add_argument("--num", type=int, default=None, help="число фраз для top")
    sp.add_argument("--period", default="monthly",
                    choices=["daily", "weekly", "monthly"], help="период для dynamics")
    sp.add_argument("--from", dest="from_date", metavar="YYYY-MM-DD",
                    help="дата начала dynamics")
    sp.add_argument("--to", dest="to_date", metavar="YYYY-MM-DD",
                    help="дата конца dynamics")
    sp.add_argument("--xlsx", help="путь к файлу (по умолч. results/<фраза>_<дата>.xlsx)")
    sp.set_defaults(func=cmd_research)

    sp = sub.add_parser("core", help="батч по файлу фраз -> агрегированное ядро")
    sp.add_argument("file", help="txt-файл, одна сид-фраза на строку")
    sp.add_argument("--clean", action="store_true",
                    help="очистить пул от нерелевантных фраз и разбить по кластерам")
    sp.add_argument("--min", dest="min_count", type=int, default=0,
                    help="мин. частотность фразы (при --clean)")
    sp.add_argument("--stop", nargs="*", default=None,
                    help="минус-слова: исключить фразы с этими словами (при --clean)")
    add_common(sp, with_num=True)
    sp.set_defaults(func=cmd_core)

    return p


def main(argv=None):
    args = build_parser().parse_args(argv)
    sys.stderr.write(f"{BRAND}\n")
    try:
        args.func(args)
    except WordstatError as e:
        sys.stderr.write(f"Ошибка: {e}\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
