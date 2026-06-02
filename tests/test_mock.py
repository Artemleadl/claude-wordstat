#!/usr/bin/env python3
"""Offline tests for wordstat.py — mock the HTTP layer, verify parsing + export."""

import csv
import io
import json
import os
import sys
import tempfile
from contextlib import redirect_stdout

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))
import wordstat  # noqa: E402

MOCK = {
    "topRequests": {
        "requestPhrase": "купить кроссовки",
        "totalCount": 123456,
        "topRequests": [
            {"phrase": "купить кроссовки мужские", "count": 45000},
            {"phrase": "купить кроссовки nike", "count": 30000},
        ],
        "associations": [{"phrase": "купить кеды", "count": 12000}],
    },
    "dynamics": {
        "dynamics": [
            {"date": "2025-11", "count": 90000, "share": 0.011},
            {"date": "2025-12", "count": 98000, "share": 0.012},
        ]
    },
    "regions": {
        "regions": [
            {"regionId": 213, "count": 50000, "share": 0.34, "affinityIndex": 145},
            {"regionId": 2, "count": 20000, "share": 0.13, "affinityIndex": 98},
        ]
    },
    "getRegionsTree": {"regions": [{"id": 225, "name": "Россия"}]},
    "userInfo": {"limit": 1000, "remaining": 873},
}

PASS = 0
FAIL = 0


def check(name, cond):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  ok  - {name}")
    else:
        FAIL += 1
        print(f"  FAIL- {name}")


def run_cmd(argv):
    """Run CLI with mocked network, capture stdout JSON (if any)."""
    buf = io.StringIO()
    with redirect_stdout(buf):
        wordstat.main(argv)
    out = buf.getvalue().strip()
    return json.loads(out) if out.startswith(("[", "{")) else out


def main():
    # Mock the network call: dispatch by method name.
    wordstat.call = lambda method, payload: MOCK[method]

    print("topRequests -> rows")
    rows = run_cmd(["top", "купить кроссовки", "--num", "200"])
    check("3 строки (2 top + 1 similar)", len(rows) == 3)
    check("seed заполнен", all(r["seed"] == "купить кроссовки" for r in rows))
    check("есть тип similar", any(r["type"] == "similar" for r in rows))
    check("count перенесён", rows[0]["count"] == 45000)
    check("seed_total = totalCount", rows[0]["seed_total"] == 123456)

    print("dynamics -> rows")
    rows = run_cmd(["dynamics", "купить елку", "--period", "monthly"])
    check("2 точки динамики", len(rows) == 2)
    check("share присутствует", rows[1]["share"] == 0.012)

    print("regions -> rows")
    rows = run_cmd(["regions", "доставка пиццы"])
    check("2 региона", len(rows) == 2)
    check("affinityIndex есть", rows[0]["affinityIndex"] == 145)

    print("quota / regions-tree (raw json)")
    info = run_cmd(["quota"])
    check("остаток квоты", info["remaining"] == 873)

    print("CSV export")
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "out.csv")
        run_cmd(["top", "купить кроссовки", "--csv", path])
        with open(path, encoding="utf-8-sig") as f:
            data = list(csv.DictReader(f))
        check("CSV: 3 строки", len(data) == 3)
        check("CSV: колонка phrase", "phrase" in data[0])

    print("XLSX export")
    try:
        import openpyxl  # noqa: F401
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "out.xlsx")
            run_cmd(["regions", "доставка пиццы", "--xlsx", path])
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.active
            check("XLSX: header + 2 строки", ws.max_row == 3)
    except ImportError:
        print("  skip- openpyxl не установлен (--xlsx опционально)")

    print("core batch")
    with tempfile.TemporaryDirectory() as d:
        seeds = os.path.join(d, "seeds.txt")
        with open(seeds, "w", encoding="utf-8") as f:
            f.write("купить кроссовки\n# комментарий\nкупить кроссовки\n")
        rows = run_cmd(["core", seeds])
        phrases = {r["phrase"] for r in rows}
        check("core: дедупликация фраз", len(phrases) == len(rows))
        check("core: собраны фразы", "купить кроссовки мужские" in phrases)

    print(f"\nИтого: {PASS} ok, {FAIL} fail")
    sys.exit(1 if FAIL else 0)


if __name__ == "__main__":
    main()
